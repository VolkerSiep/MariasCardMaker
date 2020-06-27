# -*- coding: utf-8 -*-

# stdlib modules
from os import makedirs
from os.path import dirname, join, splitext
from collections import defaultdict
from pathlib import Path

# external modules
from openpyxl import load_workbook
import wx  # assume App object exists when imported

CONFIG_SHEET_NAME = "config"
FILENAME_COL_HEADER = "filename"

class Generator:
    def __init__(self, excel_file):
        self.__excel_dir = Path(dirname(excel_file))
        self.__workbook = load_workbook(str(excel_file), data_only=True)

    def __del__(self):
        self.__workbook.close()

    def process(self, new_cards=False):
        processed = {}
        if new_cards:
            wx.LogMessage("Creating only new cards!")
        try:
            sheet = self.__workbook[CONFIG_SHEET_NAME]
        except KeyError:
            return (False, f"No Worksheet called '{CONFIG_SHEET_NAME}'")
        config = self.__read_config(sheet)

        for sheet in self.__workbook:
            if (sheet.title == CONFIG_SHEET_NAME or
                sheet.title.endswith("ignore")):
                continue
            processed[sheet.title] = self.__process_sheet(sheet, config,
                                                          new_cards)
        return processed

    def __process_sheet(self, sheet, config, new_cards):
        wx.LogMessage(f"  Processing sheet: {sheet.title}")
        # create target folder if not exists
        folder = Path(config["target_folder"]["dirname"].replace("\\", "/"))
        makedirs(self.__excel_dir / folder, exist_ok=True)

        headings = [col[0].value for col in sheet.iter_cols(max_row=0)]
        headings = [h if h else "" for h in headings]
        processed = 0
        for row in sheet.iter_rows(min_row=2):
            data = {heading: value.value
                    for heading, value in zip(headings, row[:len(headings)])
                    if (heading in config or
                        heading.lower() == FILENAME_COL_HEADER)}
            if not data["filename"]:
                break
            for page in config["pages"]:
                if self.__make_card(data, config, page, new_cards):
                    processed += 1
        return processed

    def __make_card(self, data, config, page, new_cards):
        # save picture
        target_folder = config["target_folder"]["dirname"].replace("\\", "/")
        filename = "{}_{page}{}".format(*splitext(data["filename"]),page=page)
        filename = self.__excel_dir / target_folder / filename
        if new_cards and filename.exists():
            wx.LogMessage(f"    Skipping existing card '{filename}'")
            wx.Yield()
            return False

        background = join(self.__excel_dir, config["background"]["filename"])
        bitmap = wx.Bitmap(background, wx.BITMAP_TYPE_ANY)
        dc = wx.MemoryDC(bitmap)
        # draw elements

        def add_picture(conf, file_name):
            if not file_name:
                return
            pic_name = Path(self.__excel_dir) / file_name.replace("\\", "/")
            # pic_name = join(self.__excel_dir, file_name)
            wx.LogMessage(f"    Adding picture '{pic_name}'")
            image = wx.Image(str(pic_name), wx.BITMAP_TYPE_ANY)
            max_width = conf.get("max_width", image.Width)
            max_height = conf.get("max_height", image.Height)
            scale = min([max_width / image.Width, max_height / image.Height])
            image = image.Scale(image.Width * scale,
                                image.Height * scale,
                                quality=wx.IMAGE_QUALITY_HIGH)
            y_max = dc.GetSize()[1]
            dc.DrawBitmap(wx.Bitmap(image), conf["x"], y_max - conf["y"])

        def add_text(conf, label):
            if not label:
                return
            # sc.DrawCircle()
            orgb = conf["color"]
            orgb = [int(orgb[i:i+2],16) for i in (2, 4, 6, 0)]
            dc.SetTextForeground(wx.Colour(*orgb))
            style = (wx.FONTSTYLE_ITALIC
                     if conf["italics"] else wx.FONTSTYLE_NORMAL)
            weight = (wx.FONTWEIGHT_BOLD
                     if conf["italics"] else wx.FONTWEIGHT_NORMAL)
            font = wx.Font(conf["size"], wx.FONTFAMILY_DEFAULT, style, weight,
                           conf["underline"], conf["font"]
                           )
            dc.SetFont(font)

            # align
            y_max = dc.GetSize()[1]
            # dc.DrawCircle(conf["x"], y_max - conf["y"], 2)
            x, (w, h) = conf["x"], dc.GetTextExtent(label)
            if conf["align"] == "center":
                x -= w // 2
            elif conf["align"] == "right":
                x -= w

            dc.DrawText(label, x, y_max - conf["y"] - h)

        for heading, item in data.items():
            if heading.lower() == FILENAME_COL_HEADER:
                continue
            conf = config[heading]
            if page in map(str.strip, str(conf["page"]).split(",")):
                if "picture" in heading.lower():
                    add_picture(conf, item)
                else:
                    add_text(conf, item)

        # save picture
        target_folder = config["target_folder"]["dirname"].replace("\\", "/")
        filename = "{}_{page}{}".format(*splitext(data["filename"]),page=page)
        filename = self.__excel_dir / target_folder / filename
        wx.LogMessage(f"    Saving card '{filename}'")
        wx.Yield()
        bitmap = dc.GetAsBitmap()
        bitmap.SaveFile(str(filename), wx.BITMAP_TYPE_JPEG)
        return True


    def __read_config(self, sheet):
        # find head
        for k, row in enumerate(sheet.iter_rows()):
            if row[0].value and row[0].value.upper() == "CARDSET":
                break
        else:
            return None

        # read config
        label1 = ""
        config = defaultdict(dict)
        for row in sheet.iter_rows(min_row=k + 2, max_col=3):
            label1 = row[0].value if row[0].value else label1
            if label1.lower() in ("background", "target_folder"):
                label1 = label1.lower()
            cnf = config[label1]
            label2 = row[1].value
            if not (row[0].value or row[1].value):
                break
            value = row[2].value
            if not label2:
                config[label1] = value
                continue
            label2 = label2.lower()
            if label2 == "sample":
                # cell is a font sample, extract all attributes
                align = row[2].alignment.horizontal
                cnf["align"] = "right" if align == "general" else align
                cnf["size"] = row[2].font.sz
                cnf["color"] = row[2].font.color.rgb  # orgb
                cnf["underline"] = bool(row[2].font.u)
                cnf["bold"] = bool(row[2].font.b)
                cnf["italics"] = bool(row[2].font.i)
                cnf["font"] = row[2].font.name
            else:
                cnf[label2] = value

        # determine number of pages
        pages = set()
        for item in config.values():
            pages |= set(map(str.strip, str(item.get("page", 1)).split(",")))
        config["pages"] = pages
        return config


def main():
    app = wx.App()
    file = "Cards/vocabularyAndQuestions_updated 20200621.xlsx"
    wx.Log.SetActiveTarget(wx.LogStderr())
    generator = Generator(file)
    generator.process()
    del app

if __name__ == "__main__":
    main()
